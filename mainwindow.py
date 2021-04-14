from PyQt5 import QtCore, QtGui, QtWidgets, QtMultimedia
from PyQt5.QtWidgets import QApplication, QCheckBox, QGridLayout, QTreeWidgetItem, QLabel, QLineEdit, QDialog, QMessageBox, QHBoxLayout, QGridLayout, QVBoxLayout
from PyQt5.Qt import Qt, QThread,QMutex,pyqtSignal
from ui_mainwindow import Ui_MainWindow
from PyQt5.QtGui import QPixmap, QIcon
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.Qt import *
import openpyxl
import pymysql
import os
import sys
from hashlib import md5
import random
import operator
import data_manip
import time
from ERRORS import EXECUTE_FAILURE, UPDATE_FAILURE, INSERT_FAILURE, NETWORK_ERR

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, mode, user_name, parent=None):
        super(MainWindow, self).__init__(parent)
        try:
            data_manip.check_network()
        except NETWORK_ERR as ne:
            QMessageBox.critical(self, "错误", ne.__str__())

        self.myThread = MyThread()
        self.task_running = 0
        # self.setMaximumHeight(880)
        # self.setMaximumWidth(741)
        self.setMinimumHeight(840)
        self.setMinimumWidth(560)
        self.hb1 = QVBoxLayout()
        # UI设置
        self.setupUi(self)
        QtWidgets.QApplication.setStyle('Fusion')
        # self.setWindowFlags(QtCore.Qt.FramelessWindowHint) # 设置边框是否隐藏
        self.bnt_close.setVisible(False)
        self.bnt_minimize.setVisible(False)
        self.setWindowTitle('学生辨识系统')
        self._padding = 20
        self.initDrag()
        self.setMouseTracking(True)
        self.bnt_close.setMouseTracking(True)
        self.tbw_answer = QTableWidget(self.tab)
        self.tbw_answer.hide()
        # self.lab_on_process = QLabel(self.tab)
        # self.lab_on_process.setGeometry(5, self.tabWidget.height() - 60, 100, 25)
        self.f = QtGui.QFont()
        self.f.setFamily("微软雅黑")
        self.f.setPointSize(10)
        self.lab_on_process.setFont(self.f)
        self.lab_process.setText("")
        # 管理员：mode = 0, 辅导员：mode = 1
        self.mode = mode
        self.username = user_name
        self.items = []
        self.lab_userName.setText(self.username)
        self.data = []
        self.photo_clicked = False
        self.primary_id = ''
        self.lab_saving.hide()



        # self.display_photo("xxx")
        self.student_photo.clicked.connect(self.click_to_display_info)
        self.btn_next.clicked.connect(self.on_btn_next_clicked)
        # self.btn_stop.clicked.connect(self.on_btn_stop_clicked)

        # 辅导员登录隐藏“导入辅导员”按钮
        if self.mode == 1:
            self.btn_fdy_import.hide()
            self.lab_hint.hide()
            self.cmb_select_name.hide()
        
        '''
        监听选择考核的辅导员变化
        需要仔细debug,我未确定当前逻辑的正确性
        '''
        self.implement_cmb_select_name()
       
        '''
        如果当前是管理员登录，则根据选择的辅导员进行考核
        如果是辅导员自己登录，则直接选择该辅导员自己的所有学生进行考核
        '''

        '''
        当前进度
        '''
        self.counter=0
        '''
        当前显示照片的学生id
        '''
        self.current_student_id='122'
        # self.lab_display_info.setText('')
        # self.student_photo.setStyleSheet("border-image: url('img/background.png');")
        if self.mode==0:
            self.lab_process.setText("请选择辅导员")
            self.student_photo.setEnabled(False)
            self.cmb_select_name.currentIndexChanged.connect(self.reset)
            self.cmb_select_name.currentIndexChanged.connect(self.run_select_student)
            self.student_photo.setStyleSheet("border-image: url('img/background.png');")
        else:
            # self.counter=0
            self.run_select_student()
        
        # self.student_photo.clicked.connect(self.click_to_display_info)

        # 基本属性




# ==================================组件接口==================================================
    def initDrag(self):
        # 设置鼠标跟踪判断扳机默认值
        self._move_drag = False
        self._corner_drag = False
        self._bottom_drag = False
        self._right_drag = False
        self.m_flag = False
    
    def resizeEvent(self, QResizeEvent):
    # 自定义窗口调整大小事件
        # self.tabWidget.setFixedWidth(self.width()) 
        # try:
        #     self.bnt_close.move(self.width() - self.bnt_close.width(), 0)
        # except:
        #     pass

        # try:
        #     self.bnt_minimize.move(self.width() - (self.bnt_minimize.width() + 1) * 2 + 1, 0)
        # except:
        #     pass

        # try: 
        #     self.lab_userName.move(self.width() - (self.bnt_minimize.width() + 1) * 2 - 200, 0)
        # except:
        #     pass

        # try:

            #设置所有控件的缩放规则（丑陋、愚蠢的实现方法）
        # self.lab_jump.move(int((self.tab.width() - 120) / 2), int((self.tab.height() - 30) / 2))

        self.tabWidget.resize(self.width(), self.height())
        self.lab_on_process.move(20, self.tabWidget.height() - 94)
        self.lab_saving.move(20, self.height() - 94)
        self.lab_process.move(int((self.width() - 371) / 2), 100)
        self.tab.resize(self.width(), self.height())
        self.tab_1.resize(self.width(), self.height())
        self.bnt_close.move(self.width() - self.bnt_close.width(), 0)
        self.bnt_minimize.move(self.width() - (self.bnt_minimize.width() + 1) * 2 + 1, 0)
        self.lab_userName.move(self.width() - (self.bnt_minimize.width() + 1) * 2 - 200, 10)
        self.btn_data_import.resize(int(self.width() * 200 / 741), int(50 * self.height() / 880))
        self.btn_photo_import.resize(int(self.width() * 200 / 741), int(50 * self.height() / 880))
        self.btn_delete_stud.resize(int(self.width() * 200 / 741), int(50 * self.height() / 880))
        self.btn_fdy_import.resize(int(self.width() * 200 / 741), int(50 * self.height() / 880))
        self.btn_heigth = self.btn_photo_import.height()
        self.x = int((self.width() - self.btn_delete_stud.width()) / 2)
        self.y = (self.height() - 4 * self.btn_delete_stud.height()) / 5
        self.btn_data_import.move(self.x, int(self.y))
        self.btn_photo_import.move(self.x, int(self.y * 2 + self.btn_heigth))
        self.btn_delete_stud.move(self.x, int(self.y * 3 + 2 * self.btn_heigth))
        self.btn_fdy_import.move(self.x, int(self.y * 4 + self.btn_heigth * 3))
        # self.lab_hint.move(int((self.width() - 371) / 2), int(self.height() * 50 / 880))
        # self.cmb_select_name.move(int((self.width() - 371) / 2 + 180), int(self.height() * 50 / 880))
        # self.student_photo.resize(371, int(481 * self.height() / 880))
        self.lab_hint.resize(self.tabWidget.width() * 150 / 741, self.tabWidget.height() * 20 / 880)
        self.cmb_select_name.resize(self.width() - self.lab_hint.x() - self.cmb_select_name.x(), self.tabWidget.height() * 30 / 880)
        self.lab_process.resize(self.tabWidget.width() * 350 / 741, self.tabWidget.height() * 30 / 880)
        
        # TODO 这两个按钮移动的处理可能还要调整一下
        self.btn_next.move(int((self.width() - 200) / 2), int(481 * self.height() / 880) + 250)
        # self.btn_stop.move(self.btn_next.x() + self.btn_next.width() + 150 * self.width() / 741, int(481 * self.height() / 880) + 250)
        self.btn_stop.move(int((self.width() - 200) / 2) + 200, int(481 * self.height() / 880) + 250)
        

        if not self.photo_clicked:
            # self.btn_next.move(int(self.student_photo.x() + self.student_photo.width() / 3), int(481 * self.height() / 880) + 250)
            # self.btn_stop.move(int(self.student_photo.x() + self.student_photo.width() * 2 / 3), int(481 * self.height() / 880) + 250)
            
            self.student_photo.move(int((self.width() - 371) / 2), 160)# self.student_photo.resize(self.tabWidget.width() * 371 / 741, self.tabWidget.height() * 481 / 880)
            pass
        else:
            # self.student_photo.resize(self.tabWidget.width() * 123 / 741, self.tabWidget.height() * 160 / 880)
            # self.btn_next.move(int(self.width() / 3), int(481 * self.height() / 880) + 250)
            # self.btn_stop.move(int(self.width() * 2 / 3), int(481 * self.height() / 880) + 250)
            self.tbw_answer.move(self.student_photo.x() + self.student_photo.width() + 10, 160)
            self.tbw_answer.resize(self.cmb_select_name.x() + self.cmb_select_name.width() - self.tbw_answer.x(), self.btn_next.y() - self.student_photo.y() - 30)
            self.student_photo.move(int(self.tab.width() - self.student_photo.width()) / 2, self.student_photo.y())

        try:
            # self.lab_process.move(int((self.width() - 371) / 2), 100)
            self.set_tab1_widget()
        except:
            pass
        
        try:
            self.lab_jump.move(int((self.tab.width() - 120) / 2), int((self.tab.height() - 30) / 2))
        except:
            pass

        try:
            # self.lab_display_info.move(int((self.width() - 380) / 2), self.student_photo.y() + self.student_photo.height() + 10)
            self.lab_display_info.move((self.width() - 200) // 2, self.student_photo.y() + self.student_photo.height() + 10)
        except:
            pass
        # self.set_tab1_widget()
            
        # except:
        #     pass
        # 重新调整边界范围以备实现鼠标拖放缩放窗口大小，采用三个列表生成式生成三个列表
        self._right_rect = [QPoint(x, y) for x in range(self.width() - self._padding, self.width() + 1)
                        for y in range(1, self.height() - self._padding)]
        self._bottom_rect = [QPoint(x, y) for x in range(1, self.width() - self._padding)
                        for y in range(self.height() - self._padding, self.height() + 1)]
        self._corner_rect = [QPoint(x, y) for x in range(self.width() - self._padding, self.width() + 1)
                                    for y in range(self.height() - self._padding, self.height() + 1)]
        self.drag_area = [QPoint(x, y) for x in range(150, self.lab_userName.x()) for y in range(0, 30)]

    # 关闭窗口
    @QtCore.pyqtSlot()
    def on_bnt_close_clicked(self):
        self.close()

    # 最小化窗口
    @QtCore.pyqtSlot()
    def on_bnt_minimize_clicked(self):
        self.showMinimized()

    def closeEvent(self, event):
        if self.task_running == 0:
            event.accept()
        else:
            QMessageBox.warning(self, '警告', '程序正在处理数据，请勿关闭', QMessageBox.Yes)
            event.ignore()

    # 拖动窗口
    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton and event.pos() in self.drag_area:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()
            event.accept()
            self.setCursor(QtGui.QCursor(QtCore.Qt.OpenHandCursor))

        if (event.button() == QtCore.Qt.LeftButton) and (event.pos() in self._corner_rect):
            # 鼠标左键点击右下角边界区域
            self._corner_drag = True
            event.accept()
            # self.setCursor(QtGui.QCursor(QtCore.Qt.SizeFDragCursor))

        elif (event.button() == QtCore.Qt.LeftButton) and (event.pos() in self._right_rect):
            # 鼠标左键点击右侧边界区域
            self._right_drag = True
            event.accept()
            # self.setCursor(QtGui.QCursor(QtCore.Qt.SizeHorCursor))

        elif (event.button() == QtCore.Qt.LeftButton) and (event.pos() in self._bottom_rect):
            # 鼠标左键点击下侧边界区域
            self._bottom_drag = True
            event.accept()
            # self.setCursor(QtGui.QCursor(QtCore.Qt.SizeVerCursor))

    def mouseMoveEvent(self, QMouseEvent):
        if QtCore.Qt.LeftButton and self.m_flag and QMouseEvent.pos() in self.drag_area:
            self.move(QMouseEvent.globalPos() - self.m_Position)
            QMouseEvent.accept()

        # 判断鼠标位置切换鼠标手势
        if QMouseEvent.pos() in self._corner_rect:
            self.setCursor(Qt.SizeFDiagCursor)
        elif QMouseEvent.pos() in self._bottom_rect:
            self.setCursor(Qt.SizeVerCursor)
        elif QMouseEvent.pos() in self._right_rect:
            self.setCursor(Qt.SizeHorCursor)
        else:
            self.setCursor(Qt.ArrowCursor)
        # 当鼠标左键点击不放及满足点击区域的要求后，分别实现不同的窗口调整
        # 没有定义左方和上方相关的5个方向，主要是因为实现起来不难，但是效果很差，拖放的时候窗口闪烁，再研究研究是否有更好的实现
        if Qt.LeftButton and self._right_drag:
            # 右侧调整窗口宽度
            self.resize(QMouseEvent.pos().x(), self.height())
            QMouseEvent.accept()
        elif Qt.LeftButton and self._bottom_drag:
            # 下侧调整窗口高度
            self.resize(self.width(), QMouseEvent.pos().y())
            QMouseEvent.accept()
        elif Qt.LeftButton and self._corner_drag:
            # 右下角同时调整高度和宽度
            self.resize(QMouseEvent.pos().x(), QMouseEvent.pos().y())
            QMouseEvent.accept()
        elif Qt.LeftButton and self._move_drag:
            # 标题栏拖放窗口位置
            self.move(QMouseEvent.globalPos() - self.move_DragPosition)
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False
        self.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self._move_drag = False
        self._corner_drag = False
        self._bottom_drag = False
        self._right_drag = False

    # # 导入数据
    # @QtCore.pyqtSlot()
    # def on_btn_data_import_clicked(self):
    #     filename = QtWidgets.QFileDialog.getOpenFileName(self, "打开txt文件", "", "txt文件 (*.txt)")
    #     if filename[0] != '':
    #         self.txt_path = filename[0]

    def set_tab1_widget(self):
        # self.btn_back = QPushButton(self)
        # self.lab_title = QLabel(self)
        self.btn_back.setGeometry(30, 42.5, 25, 25)
        self.lab_title.setGeometry(70, 42.5, 150, 25)
        self.btn_search.setGeometry(self.width() - 45, 42.5, 25, 25)
        self.search_box.setGeometry(self.width() - 255, 40, 200, 30)
        self.tbw_stud_info.setGeometry(30, 130, self.tabWidget.width() - 45, self.tabWidget.height() - 234)
        self.lab_not_found.setGeometry(self.search_box.x(), 80, 150, 25)
        self.btn_ok.setGeometry(self.tabWidget.width() - 115, self.tabWidget.height() - 90, 93, 28)
        self.btn_delete.setGeometry(self.tabWidget.width() - 213, self.btn_ok.y(), 93, 28)

    @QtCore.pyqtSlot()
    def on_btn_delete_stud_clicked(self):
        #"编辑学生信息按钮点击事件"
        # 隐藏界面中原来的所有控件，绘制新的控件

        self.btn_data_import.hide()
        self.btn_fdy_import.hide()
        self.btn_photo_import.hide()
        self.btn_delete_stud.hide()
        self.btn_back1 = QPushButton(self.tab_1)
        self.btn_back1.setStyleSheet("QPushButton{border-image: url(:/icon/back.png); color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;}" 
                          "QPushButton:hover{background-color: rgb(192, 192, 192);border:none;color:rgb(192, 192, 192);}")
        self.btn_back1.clicked.connect(self.quick_go_back)
        self.lab_title = QLabel(self.tab_1)
        self.btn_back1.setGeometry(30, 42.5, 25, 25)
        self.lab_title.setGeometry(70, 42.5, 150, 25)
        self.lab_title.setGeometry(70, 42.5, 150, 25)
        self.lab_title.setText("编辑学生数据")
        self.f.setPointSize(12)
        self.lab_title.setFont(self.f)
        self.f.setPointSize(10)
        self.btn_back1.show()
        self.lab_title.show()
        self.lab_jump = QLabel(self.tab_1)
        self.lab_jump.setGeometry(int((self.tab.width() - 120) / 2), int((self.tab.height() - 30) / 2), 120, 30)
        self.lab_jump.setText("正在加载...")
        self.lab_jump.show()
        # self.f.setPointSize(12)
        self.f.setPointSize(12)
        self.lab_jump.setFont(self.f)
        self.f.setPointSize(10)
        self.get_current_fdy_student()
        # try:
        #     self.get_current_fdy_student()
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     self.quick_go_back() 
        #     return
        # except EXECUTE_FAILURE as exef:
        #     QMessageBox.critical(self, "错误", exef.__str__())
        #     self.quick_go_back()
        #     return
        # self.btn_delete_stud_clicked()
    
    def quick_go_back(self):
        self.lab_jump.deleteLater()
        self.lab_title.deleteLater()
        self.btn_back1.deleteLater()
        self.btn_data_import.show()
        self.btn_photo_import.show()
        self.btn_delete_stud.show()
        if self.mode == 0:
            self.btn_fdy_import.show()

    def btn_delete_stud_clicked(self):
        #"编辑学生信息按钮点击事件"
        # 隐藏界面中原来的所有控件，绘制新的控件
       
        # self.btn_data_import.hide()
        # self.btn_fdy_import.hide()
        # self.btn_photo_import.hide()
        # self.btn_delete_stud.hide()
        # self.lab_jump = QLabel(self.tab_1)
        # self.lab_jump.setGeometry(310, 350, 120, 30)
        # self.lab_jump.setText("正在加载...")
        # # self.f.setPointSize(12)
        # self.lab_jump.setFont(self.f)
        # self.get_current_fdy_student()
        self.btn_back = QPushButton(self.tab_1)
        # self.lab_title = QLabel(self.tab_1)
        self.search_box = QLineEdit(self.tab_1)
        self.btn_search = QPushButton(self.tab_1)
        self.tbw_stud_info = QTableWidget(self.tab_1)
        self.lab_not_found = QLabel(self.tab_1)
        self.btn_ok = QPushButton(self.tab_1)
        self.btn_delete = QPushButton(self.tab_1)
        self.set_tab1_widget()

        # self.btn_back.setGeometry(30, 42.5, 25, 25)
        # self.btn_back.setPixmap(QPixmap("./images/python.jpg"))
        self.btn_back.setStyleSheet("QPushButton{border-image: url(:/icon/back.png); color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;}" 
                          "QPushButton:hover{background-color: rgb(192, 192, 192);border:none;color:rgb(192, 192, 192);}")
        
        # self.btn_delete = QPushButton(self.tab_1)
        # self.btn_delete.setGeometry(522, 790, 93, 28)
        self.btn_delete.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;")
        self.btn_delete.setText("删除")
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)

        # self.btn_ok = QPushButton(self.tab_1)
        # self.btn_ok.setGeometry(620, 790, 93, 28)
        self.btn_ok.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;")
        self.btn_ok.setText("确定")
        self.btn_ok.setFont(font)

        self.lab_title = QLabel(self.tab_1)
        self.lab_title.setGeometry(70, 42.5, 150, 25)
        self.lab_title.setText("编辑学生数据")
        font.setPointSize(12)
        self.lab_title.setFont(font)

        # self.search_box.setGeometry(470, 40, 200, 30)
        self.search_box.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;")
        self.search_box.setPlaceholderText("输入你要搜索的内容")
        font.setPointSize(10)
        self.search_box.setFont(font)

        # self.btn_search = QPushButton(self.tab_1)
        # self.btn_search.setGeometry(680, 42.5, 25, 25)
        self.btn_search.setStyleSheet("QPushButton{border-image: url(:/icon/search.png); color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 20px;  border: 2px groove gray; border-style: outset;}"
                                    "QPushButton:hover{background-color: rgb(192, 192, 192);color:rgb(192, 192, 192);}")
        
        # self.tbw_stud_info.setGeometry(30, 130, 690, 650)
        self.tbw_stud_info.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50);")
        self.tbw_stud_info.setFont(font)
        print("load_tbw_stud_info-------------------------")
        self.load_tbw_stud_info()
        print("load_tbw_stud_info-------------------------")

        # self.lab_not_found = QLabel(self.tab_1)
        # self.lab_not_found.setGeometry(470, 80, 100, 25)
        self.lab_not_found.setFont(font)

        # self.btn_ok = QPushButton(self.tab_1)
        # self.btn_ok.setGeometry(620, 790, 93, 28)
        self.btn_ok.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;")
        self.btn_ok.setText("确定")
        self.btn_ok.setFont(font)

        # self.btn_delete = QPushButton(self.tab_1)
        # self.btn_delete.setGeometry(522, 790, 93, 28)
        self.btn_delete.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50); border-radius: 10px;  border: 2px groove gray; border-style: outset;")
        self.btn_delete.setText("删除")
        self.btn_delete.setFont(font)



        self.btn_back.clicked.connect(self.go_back)
        self.btn_search.clicked.connect(self.search_support)
        self.btn_delete.clicked.connect(self.delete_stud)
        self.btn_ok.clicked.connect(self.on_btn_ok_clicked)
        self.btn_back.show()
        self.search_box.show()
        self.btn_search.show()
        self.btn_delete.show()
        self.tbw_stud_info.show()
        self.lab_title.show()
        self.lab_not_found.show()
        self.btn_ok.show()
       

        self.search_box.cursorPositionChanged.connect(self.lab_not_found.clear)



# ================================自定义方法=================================================

    def callback(self, info):
        QMessageBox.critical(self, "错误", info)
        self.quick_go_back()

    def go_back(self):  
        # print("go back-------------")
        #返回按钮点击，退出编辑界面，返回主界面。
        if self.primary_id:
            try:
                    data_manip.delete('student', 'id', self.primary_id)
                    self.primary_id = ''
            except NETWORK_ERR as ne:
                QMessageBox.critical(self, "错误", ne.__str__())
                return
            except  UPDATE_FAILURE as uf:
                QMessageBox.critical(self, "错误", uf.__str__())
                return

        if not self.check_modify():
            reply = QMessageBox.warning(self, '编辑信息', '是否保存更改？',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                student_information=[]
                temp=[]
                keys=['名字','学号','学院','专业','年级','本科/硕士','籍贯','民族','政治面貌','性格特点','辅导员姓名']
                self.data = list(self.data_modified)
                print(self.data)
                print('______________')
                print(self.data_modified)
                print('----------------')
                for i in range(len(self.data)):
                    temp=[]
                    for j in range(len(self.data[0])):
                        
                        temp.append(self.data[i][j])
                    student_information.append(temp)
                print(student_information)
                self.thread5 = MyThread()
                self.thread5.func = data_manip.add_student
                self.thread5.param = student_information
                self.lab_saving.setText("<font color='red'>   正在处理...")
                self.lab_saving.setFont(self.f)
                self.lab_saving.show()
                self.task_running += 1
                self.thread5.start()
                self.thread5.exception.connect(self.callback1)
                self.thread5.finish.connect(self.go_to_main)
        
        else:
            self.go_to_main()

    def go_to_main(self):
        # print("go back-------------")
        self.task_running -= 1
        self.lab_saving.hide()
        self.btn_back.deleteLater()
        self.btn_search.deleteLater()
        self.search_box.deleteLater()
        self.tbw_stud_info.deleteLater()
        self.btn_delete.deleteLater()
        self.lab_title.deleteLater()
        self.lab_not_found.deleteLater()
        self.btn_ok.deleteLater()
        print("go back-------------")
        self.btn_photo_import.show()
        self.btn_data_import.show()
        self.btn_delete_stud.show()
        if self.mode == 0:
            self.btn_fdy_import.show()

    def search_support(self): 
        #搜索表格内容（精准匹配，考虑修改为模糊匹配）


        if self.items:
            for item in self.items:
                item.setBackground(QBrush(QColor(50, 50, 50)))
        self.items = []

        self.tbw_stud_info.setStyleSheet("color: rgb(255, 255, 255); background-color: rgb(50, 50, 50);")
        meet = self.search_box.text()
       
        #精准匹配
        self.items = self.tbw_stud_info.findItems(meet, QtCore.Qt.MatchExactly)
        if self.items:
            for item in self.items:
                item.setBackground(QBrush(QColor(255, 255, 0)))        
        else:
            self.lab_not_found.setText("<font color='red'>   未找到相关结果")
    

    '''
    获取当前辅导员的全部学生的全部信息
    如果是管理员的话，获取所有学生的信息
    '''
    def get_current_fdy_student(self):
        thread = MyThread()
        thread.finish.connect(lambda : self.finish_load_info(thread.ret))
        thread.exception.connect(self.callback)
        if self.mode==1:
            
            thread.func = data_manip.get_student_info
            thread.param = self.username[3:]
            thread.start()
            # self.data = thread.ret
            # ret=data_manip.get_student_info(self.username[3:])
            # if thread.ret[0]==0:
            #     reply = QMessageBox.information(self,
            #                         "提示",  
            #                         "该辅导员没有学生",  
            #                         QMessageBox.Yes | QMessageBox.No)
            # self.data=thread.ret
        else:

            thread.func = data_manip.get_all_information
            thread.param = {}
            thread.start()
            # self.data = thread.ret
            # try:
            #     thread.start()
            # except NETWORK_ERR as ne:
            #     QMessageBox.critical(self, "错误", ne.__str__())
            # except EXECUTE_FAILURE as exef:
            #     QMessageBox.critical(self, "错误", exef.__str__())

            # self.data=data_manip.get_all_information()
        
    def finish_load_info(self, ret):
        self.lab_jump.deleteLater()
        self.btn_back1.deleteLater()
        self.lab_title.deleteLater()
        if self.mode == 1:
            if ret[0] == 0:
                reply = QMessageBox.information(self,
                                    "提示",  
                                    "该辅导员没有学生",  
                                    QMessageBox.Yes | QMessageBox.No)
            self.data = ret

        else:
            self.data = ret
            pass

        self.btn_delete_stud_clicked()
    #加载“编辑学生信息”中的表格内容
    def load_tbw_stud_info(self):
        if self.data:
            print("this is the data", self.data)
            self.header = ['名字','学号','学院','专业','年级','本科/硕士','籍贯','民族','政治面貌','性格特点','辅导员姓名']
            self.tbw_stud_info.setRowCount(len(self.data))
            self.tbw_stud_info.setColumnCount(len(self.header))
            self.tbw_stud_info.setHorizontalHeaderLabels(self.header)
            self.tbw_stud_info.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)

            for row in range(len(self.data)):
                col = 0
                for i in range(0,11):
                    temp = str(self.data[row][i])
                    print(temp)
                    item = QtWidgets.QTableWidgetItem(temp)
                    item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                    self.tbw_stud_info.setItem(row, col, item)
                    col += 1
        else:
            print("no dataad")

    #"编辑学生信息"中的删除按钮逻辑
    def delete_stud(self):
        item = self.tbw_stud_info.selectedItems()
        self.primary_id = ''
        if item:
            row_selected = item[0].row()
            self.primary_id = self.tbw_stud_info.item(row_selected, 1).text()
            self.tbw_stud_info.removeRow(row_selected)
            

    #"确定"按钮和"返回"按钮的工具函数
    def check_modify(self):
        header = []
        for i in range(self.tbw_stud_info.columnCount()):
            header.append(self.tbw_stud_info.horizontalHeaderItem(i).text())
 
        self.data_modified = []
        row = self.tbw_stud_info.rowCount()

        for item in range(row):
            dic = {}
            for col in range(len(header)):
                if self.tbw_stud_info.item(item, col) != None:
                    record = self.tbw_stud_info.item(item, col).text()
                    dic[header[col]] = record
            self.data_modified.append(dic)
        data2=[]
        for i in range(len(self.data_modified)):
            data=[]
            for key in ['名字','学号','学院','专业','年级','本科/硕士','籍贯','民族','政治面貌','性格特点','辅导员姓名']:
                data.append(self.data_modified[i][key])
            data=tuple(data)
            data2.append(data)
        self.data_modified=tuple(data2)    
        print("data:")
        print(self.data)
        print('=' * 20)
        print(self.data_modified)
        print(operator.eq(self.data, self.data_modified))
        return operator.eq(self.data, self.data_modified)

    #"编辑学生信息"确定按钮逻辑
    def on_btn_ok_clicked(self):
        # reply = QMessageBox.information(self, "提示", "删除成功", QMessageBox.Yes | QMessageBox.No)
        # if not self.check_modify():
        #     self.data = list(self.data_modified)
        self.go_back()
    

    '''
    显示照片
    1.选择学生
    2.显示图片
    3.点击显示信息
    '''
    def reset(self):
        self.photo_clicked = False
        self.student_photo.setStyleSheet("border-image: url('img/background.png');")
        self.lab_process.setText("请选择辅导员")
        self.student_photo.move(int((self.tabWidget.width() - 371) / 2), 160)
        # if self.photo_clicked == True:
        #     self.photo_clicked = False
        #     self.tbw_answer.deleteLater()
            # self.student_photo.move(int((self.tabWidget.width() - 371) / 2), 160)
        self.student_photo.resize(371, 481)
        self.student_photo.setEnabled(False)
        self.lab_display_info.hide()
        self.counter=0
    def callback2(self, info):
        self.student_photo.setText("加载失败，网络连接中断")
        self.student_photo.setFont(self.f)

    def run_select_student(self):
        # print(self.cmb_select_name.cu)
        if self.mode==0:
            self.current_fdy_name=self.cmb_select_name.currentText()
            self.lab_display_info.setText('')
        else:
            self.current_fdy_name=self.username[3:]
            '''
            选择当前辅导员的所有的学生
            '''

        if self.mode == 0:
            if self.cmb_select_name.currentIndex() != 0:
                self.student_photo.setStyleSheet("background-color: rgb(50, 50, 50); color: rgb(255, 255, 255);")
                self.student_photo.setText("正在加载，请稍候...")
                self.student_photo.setEnabled(False)
                # try:
                #     data_manip.check_network()
                # except NETWORK_ERR as ne:
                #     QMessageBox.critical(self, "错误", ne.__str__())
                #     return

                # self.select_student_thread = MyThread(self.select_student, {})
                self.myThread.func = self.select_student
                self.myThread.param = self.current_fdy_name
                self.myThread.start()
                self.myThread.exception.connect(self.callback2)
                self.myThread.finish.connect(lambda : self.load_stud_photo_finish(self.myThread.ret))

            else:
                self.reset()
        else:
            self.student_photo.setStyleSheet("background-color: rgb(50, 50, 50); color: rgb(255, 255, 255);")
            self.student_photo.setText("正在加载，请稍候...")
            self.student_photo.setEnabled(False)
            # try:
            #     data_manip.check_network()
            # except NETWORK_ERR as ne:
            #     QMessageBox.critical(self, "错误", ne.__str__())
            #     return

            # self.select_student_thread = MyThread(self.select_student, {})
            self.myThread.func = self.select_student
            self.myThread.param = self.current_fdy_name
            self.myThread.start()
            self.myThread.exception.connect(self.callback2)
            self.myThread.finish.connect(lambda : self.load_stud_photo_finish(self.myThread.ret))


    def select_student(self, current_fdy_name):
        # self.student_photo.setStyleSheet("border-image: url('img/background.png');")

        # if self.mode==0:
        #     self.current_fdy_name=self.cmb_select_name.currentText()
        #     self.lab_display_info.setText('')
        # else:
        #     self.current_fdy_name=self.username[3:]
        #     '''
        #     选择当前辅导员的所有的学生
        #     '''
        return data_manip.get_photo(current_fdy_name)
        # try:
        #     temp=data_manip.get_photo(self.current_fdy_name)
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return
        # except EXECUTE_FAILURE as exef:
        #     QMessageBox.critical(self, "错误", exef.__str__())

    def load_stud_photo_finish(self, temp):
        self.student_photo.setText('')
        if temp[0]==0:
            # reply = QMessageBox.information(self,
            #                         "提示",  
            #                         "该辅导员没有学生",  
            #                         QMessageBox.Yes | QMessageBox.No)
            self.student_photo.setText("查询失败")
            self.lab_process.setText('当前进度: 0|0')
            self.student_photo.setFont(self.f)

        else:
            self.student_photo.setText("")
            results=temp[1]
            if not results:
                self.student_photo.setText("当前辅导员暂无学生照片")
                self.lab_process.setText('当前进度: 0|0')
                self.student_photo.setFont(self.f)

            else:
                student_list=[]

                for i in range(0,len(results)):
                        student_list.append(results[i][0])
                        path='./temp'
                        if not os.path.exists(path):
                            os.makedirs(path)
                        files= os.listdir(path) #得到文件夹下的所有文件名称
                        flag=False
                        for file in files: #遍历文件
                            if os.path.splitext(file)[0] == results[i][0]:
                                flag=True
                        if flag==False:
                            self.download_photo(results[i][1],results[i][0])
                
                if self.mode==0:
                    if len(student_list)!=0:
                        if self.counter==0:
                        #随机打乱列表
                            random.shuffle(student_list)
                        if self.counter>=len(student_list):
                            self.counter=0
                        self.current_student_id=student_list[self.counter]
                        self.lab_process.setText(('当前进度: %s'%str(self.counter+1))+' | %s'%len(student_list))
                        self.counter+=1
                        self.display_photo(self.current_student_id)
                    else:
                        return
                else:
                    if self.counter>=len(student_list):
                        self.counter=0
                
                    self.current_student_id=student_list[self.counter]
                    self.lab_process.setText(('当前进度: %s'%str(self.counter+1))+'| %s'%len(student_list))
                    self.counter+=1
                    self.display_photo(self.current_student_id)
           


    def display_photo(self, id):
        self.student_photo.setEnabled(True)
        path='temp/'+id+'.jpg'
        self.student_photo.setStyleSheet("border-image: url(%s);"%path)


    '''
    更改后的点击学生照片查询信息
    '''
    def callback3(self, info):
        QMessageBox.critical(self, "错误", info)
        return

    def click_to_display_info(self):
        if not self.photo_clicked:
            self.photo_clicked = True
            thread4 = MyThread()
            thread4.func = data_manip.get_student_info1
            thread4.param = self.current_student_id
            thread4.start()
            thread4.finish.connect(lambda : self.load_stud_photo_info_finish(thread4.ret))
            thread4.exception.connect(self.callback3)
        # try:
        #     results=data_manip.get_student_info1(self.current_student_id)
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return
        # except EXECUTE_FAILURE as exef:
        #     QMessageBox.critical(self, "错误", exef.__str__())
        #     return

    def load_stud_photo_info_finish(self, results):
        self.student_photo.resize(self.student_photo.width() / 2, self.student_photo.height() / 2)
        self.student_photo.move(int(self.tab.width() - self.student_photo.width()) / 2, self.student_photo.y())
        
        self.lab_display_info.setGeometry(180, self.student_photo.y() + self.student_photo.height() + 10 ,380,300)
        self.lab_display_info.move((self.width() - 200) // 2, self.student_photo.y() + self.student_photo.height() + 10)

        if len(results)==0:
            return
        information='名字: '+results[0][0]+'\n'+'学号: '+results[0][1]+'\n'+'学院: '+results[0][2]+'\n'+'专业: '+results[0][3]+'\n'+'年级: '+results[0][4]+'\n'+'本科\硕士: '+results[0][5]+'\n'+'籍贯: '+results[0][6]+'\n'+'民族: '+results[0][7]+'\n'+'政治面貌: '+results[0][8]+'\n'+'性格特点: '+results[0][9]
        self.lab_display_info.setText(information)
        self.lab_display_info.show()
   
       
    
    
    def download_photo(self,image,id):
        if not os.path.exists('./temp'):
            os.makedirs('./temp')
        with open('./temp/' + id+'.jpg','wb') as f:
            f.write(image)    
        print('./temp/' + id +'.jpg')

    '''
    显示下一张图片
    '''
    @QtCore.pyqtSlot()
    def on_btn_next_clicked(self):
        self.photo_clicked = False
        self.student_photo.setEnabled(True)
        # if self.photo_clicked == True:
        #     self.photo_clicked = False
        #     self.tbw_answer.deleteLater()
        self.student_photo.move(int((self.tabWidget.width() - 371) / 2), 160)
        self.student_photo.resize(371, 481)
        self.lab_display_info.setText('')
        self.lab_display_info.hide()
        self.run_select_student()
        


    '''
    停止
    '''
    @QtCore.pyqtSlot()
    def on_btn_stop_clicked(self):
        self.current_student_id='456'
        # if self.photo_clicked == True:
        #     self.photo_clicked = False
        #     self.tbw_answer.deleteLater()
        self.btn_stop.setStyleSheet("border-image:url('img/stop2.png')")
        reply = QMessageBox.question(self,
                                    "提示",  
                                    "当前已经答题%s"%str(self.counter)+',确定要将数据清零吗？',  
                                    QMessageBox.Yes | QMessageBox.No)
        if reply== QMessageBox.Yes:
            self.counter=0
            self.student_photo.setStyleSheet("border-image: url('img/background.png');")
            self.lab_process.setText('')
        self.btn_stop.setStyleSheet("border-image:url('img/stop.png')")
        self.student_photo.resize(371,481)
        self.student_photo.move(int((self.tabWidget.width() - 371) / 2), 160)

    def callback1(self, info):
        QMessageBox.critical(self, "错误", info)
        self.task_running -= 1
        self.lab_on_process.setText("")


     # 导入学生数据
    @QtCore.pyqtSlot()
    def on_btn_data_import_clicked(self):
        self.xlsx_path=" "
        filename = QtWidgets.QFileDialog.getOpenFileName(self, "打开包含学生信息的xlsx文件", "", "xlsx文件 (*.xlsx)")
        if filename[0] != '':
            self.xlsx_path= filename[0]
        else:
            return
        #读取文件
        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        sheets = wb.worksheets
        sheet=sheets[0]
        student_info=[]
        for i in range(2,sheet.max_row+1):
            student=[]
            student.append(sheet.cell(row=i,column=2).value)
            student.append(sheet.cell(row=i,column=1).value)
            student.append(sheet.cell(row=i,column=3).value)
            student.append(sheet.cell(row=i,column=4).value)
            student.append(sheet.cell(row=i,column=5).value)
            student.append(sheet.cell(row=i,column=6).value)
            student.append(sheet.cell(row=i,column=7).value)
            student.append(sheet.cell(row=i,column=8).value)
            student.append(sheet.cell(row=i,column=9).value)
            student.append(sheet.cell(row=i,column=10).value)
            student.append(sheet.cell(row=i,column=11).value)
            student_info.append(student)

        # try:
        #     data_manip.check_network()
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return

        self.task_running += 1
        self.lab_on_process.setText("<font color='red'>   正在处理...")
        self.lab_on_process.show()
        thread_1 = MyThread()
        thread_1.func = self.btn_data_import_clicked
        thread_1.param = student_info
        info_success = "恭喜你，学生信息已经成功添加"
        info_fail = "部分学生未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！"
        # self.thread_1.win = self
        thread_1.start()
        thread_1.exception.connect(self.callback1)
        # print("task number before: ", self.task_running)
        thread_1.finish.connect(lambda : self.one_task_finished(thread_1.ret, info_success, info_fail))
        # print("task number after: ", self.task_running)

        # if mythread.ret==0:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "部分学生未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！",  
        #                             QMessageBox.Yes | QMessageBox.No)
        # else:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "恭喜你，学生信息已经成功添加",  
        #                             QMessageBox.Yes | QMessageBox.No)

        # if self.task_running == 0:
        #     self.lab_on_process.setText("")

        return

    
    def one_task_finished(self, ret, s_info, f_info):
        if self.task_running == 0:
            print("here")
            return
        else:
            print("one tast finish")
            self.task_running -= 1

        if self.task_running == 0:
            self.lab_on_process.setText("")

        if ret == 1:
            QMessageBox.information(self, "提示", s_info, QMessageBox.Yes | QMessageBox.No)
        else:
            QMessageBox.information(self, "提示", s_info, QMessageBox.Yes | QMessageBox.No)
        # print("after finish", self.task_running)
    
    def btn_data_import_clicked(self, student_info):
        return data_manip.add_student(student_info)
        # self.xlsx_path=" "
        # filename = QtWidgets.QFileDialog.getOpenFileName(self, "打开包含学生信息的xlsx文件", "", "xlsx文件 (*.xlsx)")
        # if filename[0] != '':
        #     self.xlsx_path= filename[0]
        # else:
        #     return
        # #读取文件
        # wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        # sheets = wb.worksheets
        # sheet=sheets[0]
        # student_info=[]
        # for i in range(2,sheet.max_row+1):
        #     student=[]
        #     student.append(sheet.cell(row=i,column=2).value)
        #     student.append(sheet.cell(row=i,column=1).value)
        #     student.append(sheet.cell(row=i,column=3).value)
        #     student.append(sheet.cell(row=i,column=4).value)
        #     student.append(sheet.cell(row=i,column=5).value)
        #     student.append(sheet.cell(row=i,column=6).value)
        #     student.append(sheet.cell(row=i,column=7).value)
        #     student.append(sheet.cell(row=i,column=8).value)
        #     student.append(sheet.cell(row=i,column=9).value)
        #     student.append(sheet.cell(row=i,column=10).value)
        #     student.append(sheet.cell(row=i,column=11).value)
        #     student_info.append(student)
        # try:
        #     ret = data_manip.add_student(student_info)
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return
        # if ret == 0:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "部分学生未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！",  
        #                             QMessageBox.Yes | QMessageBox.No)
        # else:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "恭喜你，学生信息已经成功添加",  
        #                             QMessageBox.Yes | QMessageBox.No)

        # self.one_task_finished()
        # if self.task_running == 0:
        #     self.lab_on_process.setText("")

        # return

            
         
    #导入辅导员
    @QtCore.pyqtSlot()
    def on_btn_fdy_import_clicked(self):
        self.xlsx_path=" "
        filename = QtWidgets.QFileDialog.getOpenFileName(self, "打开包含辅导员姓名和密码的xlsx文件", "", "xlsx文件 (*.xlsx)")
        if filename[0] != '':
            self.xlsx_path= filename[0]
        else:
            return
        #读取文件
        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        sheets = wb.worksheets
        sheet=sheets[0]
        fdy_info=[]
        for i in range(2,sheet.max_row+1):
            name=sheet.cell(row=i,column=1).value
            password=sheet.cell(row=i,column=2).value
            temp=[]
            temp.append(name)
            temp.append(password)
            fdy_info.append(temp)

        # try:
        #     data_manip.check_network()
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return

        self.task_running += 1
        self.lab_on_process.setText("<font color='red'>   正在处理...")
        self.lab_on_process.show()
        thread_2 = MyThread()
        thread_2.func = self.btn_fdy_import_clicked
        thread_2.param = fdy_info
        info_success = "恭喜你，辅导员账号已经成功添加！"
        info_fail = "部分辅导员账号未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！"
        # self.thread_2.win = self
        thread_2.start()
        thread_2.finish.connect(lambda : self.one_task_finished(thread_2.ret, info_success, info_fail))
        thread_2.exception.connect(self.callback1)
        # print("Ret outside", thread.ret)

        # if thread.ret==0:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "部分辅导员账号未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！",  
        #                             QMessageBox.Yes | QMessageBox.No)
        # else:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "恭喜你，辅导员账号已经成功添加！",  
        #                             QMessageBox.Yes | QMessageBox.No)

        # if self.task_running == 0:
        #     self.lab_on_process.setText("")

    #导入辅导员
    @QtCore.pyqtSlot()
    def btn_fdy_import_clicked(self, fdy_info):
        return data_manip.add_fdy(fdy_info)
        # self.xlsx_path=" "
        # filename = QtWidgets.QFileDialog.getOpenFileName(self, "打开包含辅导员姓名和密码的xlsx文件", "", "xlsx文件 (*.xlsx)")
        # if filename[0] != '':
        #     self.xlsx_path= filename[0]
        # else:
        #     return
        # #读取文件
        # wb = openpyxl.load_workbook(self.xlsx_path, read_only=True)
        # sheets = wb.worksheets
        # sheet=sheets[0]
        # fdy_info=[]
        # for i in range(2,sheet.max_row+1):
        #     name=sheet.cell(row=i,column=1).value
        #     password=sheet.cell(row=i,column=2).value
        #     temp=[]
        #     temp.append(name)
        #     temp.append(password)
        #     fdy_info.append(temp)
        # try:
        #     return data_manip.add_fdy(fdy_info)
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return
        # if ret==0:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "部分辅导员账号未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！",  
        #                             QMessageBox.Yes | QMessageBox.No)
        # else:
        #     reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "恭喜你，辅导员账号已经成功添加！",  
        #                             QMessageBox.Yes | QMessageBox.No)


    #导入学生照片
    @QtCore.pyqtSlot()
    def on_btn_photo_import_clicked(self):
        #存放照片的文件夹路径
        self.photo_path=" "
        filename = QtWidgets.QFileDialog.getExistingDirectory(self, "打开包含学生照片的文件夹", "")
        if filename!= '':
            self.photo_path= filename
            print(self.photo_path)
        else:
            return
        files= os.listdir(self.photo_path) #得到文件夹下的所有文件名称
        for file in files: 
            path=os.path.join(self.photo_path,file)
            student_id=os.path.splitext(file)[-2]
            isvalid=os.path.splitext(file)[-1]
            if isvalid!='.jpg'and isvalid!='.png'and isvalid!='.jpeg'and isvalid!='.JPG'and isvalid!='.PNG'and isvalid!='.JPEG':
                continue 
            fp=open(path,'rb')
            img=fp.read()
            fp.close()

        # try:
        #     data_manip.check_network()
        # except NETWORK_ERR as ne:
        #     QMessageBox.critical(self, "错误", ne.__str__())
        #     return

        self.task_running += 1
        self.lab_on_process.setText("<font color='red'>   正在处理...")
        self.lab_on_process.show()
        thread_3 = MyThread()
        thread_3.func = self.btn_photo_import_clicked
        dic = {"student_id" : student_id, "photo" : img}
        thread_3.param = dic
        info_success = "恭喜你，学生照片已经成功添加！"
        info_fail = "学生%s的照片未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！"%student_id
        # self.thread_3.win = self
        thread_3.start()
        thread_3.finish.connect(lambda : self.one_task_finished(thread_3.ret, info_success, info_fail))
        thread_3.exception.connect(self.callback1)
        # if ret==0:
        #         reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "学生%s的照片未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！"%student_id,  
        #                             QMessageBox.Yes | QMessageBox.No)
            
        # reply = QMessageBox.information(self,
        #                     "提示",  
        #                     "恭喜你，学生照片已经成功添加！",  
        #                     QMessageBox.Yes | QMessageBox.No)

        # if self.task_running == 0:
        #     self.lab_on_process.setText("")
         
    #导入学生照片
    @QtCore.pyqtSlot()
    def btn_photo_import_clicked(self, param):
        return data_manip.add_photo(param)
        # #存放照片的文件夹路径
        # self.photo_path=" "
        # filename = QtWidgets.QFileDialog.getExistingDirectory(self, "打开包含学生照片的文件夹", "")
        # if filename!= '':
        #     self.photo_path= filename
        #     print(self.photo_path)
        # else:
        #     return
        # files= os.listdir(self.photo_path) #得到文件夹下的所有文件名称
        # for file in files: 
        #     path=os.path.join(self.photo_path,file)
        #     student_id=os.path.splitext(file)[-2]
        #     isvalid=os.path.splitext(file)[-1]
        #     if isvalid!='.jpg'and isvalid!='.png'and isvalid!='.jpeg'and isvalid!='.JPG'and isvalid!='.PNG'and isvalid!='.JPEG':
        #         continue 
        #     fp=open(path,'rb')
        #     img=fp.read()
        #     fp.close()
            # try:
            #     return data_manip.add_photo(param)
            # except NETWORK_ERR as ne:
            #     QMessageBox.critical(self, "错误", ne.__str__())
            #     return
        #     if ret==0:
        #         reply = QMessageBox.information(self,
        #                             "提示",  
        #                             "学生%s的照片未能成功添加，你可以重新添加该文件，已经插入的学生信息会被覆盖！"%student_id,  
        #                             QMessageBox.Yes | QMessageBox.No)
            
        # reply = QMessageBox.information(self,
        #                     "提示",  
        #                     "恭喜你，学生照片已经成功添加！",  
        #                     QMessageBox.Yes | QMessageBox.No)



    '''
    implementation of the Qcombox (cmb_select_name)
    '''

    def implement_cmb_select_name(self):
        allName=[]
        try:
            results=data_manip.get_fdy()
        except NETWORK_ERR as ne:
            QMessageBox.critical(self, "错误", ne.__str__())
            return
        except EXECUTE_FAILURE as exef:
            QMessageBox.critical(self, "错误", exef.__str__())
            return
        for i in range(len(results)):
                allName.append(results[i][0])

        if len(allName)!=0:
            for name in allName:
               
                self.cmb_select_name.addItem(name)
            '''
            当前选定的辅导员姓名
            '''
        else:
            return
    

   


        

#     #展示学生相片
#     def display_photo(self, path):
#         """
#         param: path: 接收照片的相对路径
#         """
#         #测试用
#         self.student_photo.setStyleSheet("border-image: url(%s);" % path)

#     #点击学生相片后显示详细信息的逻辑
#     def click_to_display_info(self):
#         #点击相片后显示学生信息
#         # self.lab_display_info.setText("显示学生信息")
#         if not self.photo_clicked:
#             self.photo_clicked = True
#             self.student_photo.resize(int(self.student_photo.width() / 3), int(self.student_photo.height() / 3))
#             self.student_photo.move(self.lab_hint.x(), self.student_photo.y())
#             self.tbw_answer = QTableWidget(self.tab)
#             self.tbw_answer.setGeometry(self.student_photo.x() + self.student_photo.width() + 10, self.student_photo.y(), self.width() - self.student_photo.x() - self.student_photo.width() - 40, self.btn_next.y() - self.student_photo.y() - 30)
#             self.tbw_answer.show()
#             #接下去设置表格内容（国杰已实现）
    
#     #"下一张"按钮的逻辑
#     # def on_btn_next_clicked(self):
#     #     if self.photo_clicked:
#     #         self.photo_clicked = False
#     #         self.student_photo.setGeometry(190, 160, 371, 481)
#     #         self.tbw_answer.deleteLater()
        
#         #展示下一张照片（国杰已实现）

#     def on_btn_stop_clicked(self):
#         if self.photo_clicked:
#             self.photo_clicked = False
#             self.student_photo.setGeometry(190, 160, 371, 481)
#             self.tbw_answer.deleteLater()



# if __name__=="__main__":
#     app = QtWidgets.QApplication(sys.argv)
#     db = pymysql.connect(host="gz-cynosdbmysql-grp-0965sb99.sql.tencentcdb.com", user="root", password="A1b1c1d1", db="test", port=25462,charset='utf8')
#     cr = db.cursor(cursor = pymysql.cursors.DictCursor)
#     sql = "SELECT * FROM `Data`"
#     cr.execute(sql)
#     data = cr.fetchall()
#     win = MainWindow(0, '黄明')
#     win.data = data
#     win.show()
#     # win.data = data
#     sys.exit(app.exec_())

class MyThread(QThread):
    # signal = pyqtSignal(str)
    finish = pyqtSignal()
    exception = pyqtSignal(str)
    # ret = pyqtSignal(int)
    def __init__(self):
        super(MyThread, self).__init__()
        self.func = None
        self.ret = None
        self.info_success = ''
        self.info_fail = ''
        self.param = {}
        # self.win = None

    def run(self):
        if not self.param:
            # time.sleep(5)
            try:
                self.ret = self.func()
                time.sleep(10)
            except NETWORK_ERR as ne:
                self.exception.emit(ne.__str__())
                return 
            except EXECUTE_FAILURE as exef:
                self.exception.emit(exef.__str__())
                return
                
            
        else:
            try:
                self.ret = self.func(self.param)
                time.sleep(10)
            except NETWORK_ERR as ne:
                self.exception.emit(ne.__str__())
                self.terminate
                return
            except EXECUTE_FAILURE as exef:
                self.exception.emit(exef.__str__())
                self.terminate
                return
            # time.sleep(10)
        
        # if self.ret == 1:
        #     QMessageBox.information(self, "提示", self.info_success, QMessageBox.Yes)
        # else:
        #     QMessageBox.information(self, "提示", self.info_fail, QMessageBox.Yes)

        self.finish.emit()
        print("finish")
        # print("ret = ", self.ret)

        # if self.res != None:
        #     self.ret.emit(self.res)
